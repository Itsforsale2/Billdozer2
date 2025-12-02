from openpyxl import load_workbook

excel_path = r"C:\Python\Python Projects\Billing summary\Billing_Summary\Month\Billing_summary.xlsx"

wb = load_workbook(excel_path)

print("Sheets:", wb.sheetnames)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print("\n=== Sheet:", sheet, "===")
    for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
        print(row)
