import os
import shutil
import fitz  # PyMuPDF
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook

# ============================================================
# VENDOR PARSER ALIASES  (folder name → parser module)
# ============================================================
VENDOR_ALIASES = {
    "knife river": "knife_river_parser",
    "core & main": "core_main_parser",
    # add more vendors here manually as you create parsers:
    # "farwest": "farwest_parser",
    # "sunbelt": "sunbelt_parser",
}




# ------------------------------------------------------------
# CONSTANTS / PATHS
# ------------------------------------------------------------

# Root folder (same root as this file)
ROOT_DIR = os.path.dirname(os.path.abspath(__file__))

# Job list file (user-provided path)
JOB_LIST_PATH = os.path.join(ROOT_DIR, "helperfiles","jobs.txt")

# Billing summary root (inside ROOT_DIR)
BILLING_SUMMARY_ROOT = os.path.join(ROOT_DIR, "Billing_Summary")


def sanitize_sheet_title(title: str) -> str:
    """
    Sanitize an Excel sheet name:
    - Remove forbidden characters: : \ / ? * [ ]
    - Trim to 31 characters.
    - Fallback to 'Sheet1' if empty.
    """
    forbidden = [":", "\\", "/", "?", "*", "[", "]"]
    cleaned = "".join(ch for ch in title if ch not in forbidden)
    cleaned = cleaned.strip()
    if not cleaned:
        cleaned = "Sheet1"
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned


def safe_move(src: str, dst_folder: str) -> str:
    """
    Move src file into dst_folder.
    If a file with the same name exists, append _1, _2, ... before the extension.
    Returns final destination path.
    """
    os.makedirs(dst_folder, exist_ok=True)
    base = os.path.basename(src)
    name, ext = os.path.splitext(base)
    candidate = os.path.join(dst_folder, base)
    counter = 1
    while os.path.exists(candidate):
        candidate = os.path.join(dst_folder, f"{name}_{counter}{ext}")
        counter += 1
    shutil.move(src, candidate)
    return candidate


class InvoiceSorterUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Invoice Sorting System")
        self.geometry("1700x900")
        self.configure(bg="#1e1e1e")

        self.zoom_level = 1.5

        # state
        self.invoices = []              # list of pdf paths currently shown in left list
        self.current_folder = None      # current folder backing left list
        self.pdf_document = None        # fitz.Document for center preview
        self.displayed_page = 0
        self.vendor_module_name = None  # e.g., "knife_river_parser"
        self.last_parsed_invoice = None # dict for currently selected invoice

        self.jobnames = []              # list of job names loaded from file
        self.job_selection_auto = False # True when selection is auto-highlighted

        # Per-invoice info: path -> {"parsed": dict|None, "jobname": str|None, "staged": bool}
        self.invoice_info = {}

        # MAIN PANED LAYOUT (LEFT • CENTER • RIGHT)
        self.paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        self.paned.pack(fill="both", expand=True)

        self.left_frame = tk.Frame(self, bg="#252526")
        self.center_frame = tk.Frame(self, bg="#111111")
        self.right_frame = tk.Frame(self, bg="#252526")

        self.paned.add(self.left_frame, weight=1)
        self.paned.add(self.center_frame, weight=4)
        self.paned.add(self.right_frame, weight=2)

        self.build_left_panel()
        self.build_center_panel()
        self.build_right_panel()

        # load job names from file
        self.load_jobnames()

        # ====================== PATCH START (CTRL+R binding) ======================
        self.bind_all("<Control-r>", self.secret_auto_stage)

        # ====================== PATCH END ======================



    # =========================================================
    # LEFT PANEL
    # =========================================================
    def build_left_panel(self):
        frame = self.left_frame
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(2, weight=1)

        # Load folder button
        btn = tk.Button(
            frame,
            text="Load Invoice Folder",
            bg="#3a3d41",
            fg="white",
            command=self.load_folder,
        )
        btn.grid(row=0, column=0, sticky="ew", pady=10, padx=10)

        # Vendor label
        self.vendor_label = tk.Label(
            frame,
            text="Vendor: (none selected)",
            fg="#9cdcfe",
            bg="#252526",
            anchor="w",
        )
        self.vendor_label.grid(row=1, column=0, sticky="ew", padx=10)

        # List of PDFs
        self.invoice_list = tk.Listbox(
            frame,
            bg="#1e1e1e",
            fg="white",
            selectmode=tk.SINGLE,
            exportselection=False,  # allow independent selection vs job list
        )
        self.invoice_list.grid(row=2, column=0, sticky="nsew", padx=10, pady=10)
        self.invoice_list.bind("<<ListboxSelect>>", self.on_invoice_select)

        # PROCESS INVOICES BUTTON
        process_btn = tk.Button(
            frame,
            text="Process Invoices",
            bg="#007acc",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            command=self.process_invoices,
        )
        process_btn.grid(row=3, column=0, sticky="ew", padx=10, pady=(5, 20))

    # =========================================================
    # CENTER PANEL (PDF Display + Zoom)
    # =========================================================
    def build_center_panel(self):
        frame = self.center_frame
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        # Scrollable canvas
        self.pdf_canvas = tk.Canvas(frame, bg="#111111")
        self.pdf_canvas.grid(row=0, column=0, sticky="nsew")

        # Add scrollbars
        vbar = tk.Scrollbar(frame, orient=tk.VERTICAL, command=self.pdf_canvas.yview)
        hbar = tk.Scrollbar(frame, orient=tk.HORIZONTAL, command=self.pdf_canvas.xview)
        self.pdf_canvas.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)
        vbar.grid(row=0, column=1, sticky="ns")
        hbar.grid(row=1, column=0, sticky="ew")

        # Zoom with Ctrl + MouseWheel
        self.pdf_canvas.bind("<Control-MouseWheel>", self.on_zoom)

    # =========================================================
    # RIGHT PANEL (PARSED TOP • JOB LIST + BUTTONS BOTTOM)
    # =========================================================
    def build_right_panel(self):
        frame = self.right_frame
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=2)
        frame.rowconfigure(3, weight=3)

        parsed_lbl = tk.Label(
            frame,
            text="Parsed Invoice Data",
            fg="#9cdcfe",
            bg="#252526",
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        )
        parsed_lbl.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 5))

        self.parsed_output = tk.Text(
            frame,
            bg="#1e1e1e",
            fg="white",
            height=10,
            wrap="word",
        )
        self.parsed_output.grid(row=1, column=0, sticky="nsew", padx=10)

        job_lbl = tk.Label(
            frame,
            text="Job List",
            fg="#9cdcfe",
            bg="#252526",
            font=("Segoe UI", 12, "bold"),
            anchor="w",
        )
        job_lbl.grid(row=2, column=0, sticky="ew", padx=10, pady=(10, 5))

        # job list + scrollbar
        job_frame = tk.Frame(frame, bg="#252526")
        job_frame.grid(row=3, column=0, sticky="nsew", padx=10)
        job_frame.rowconfigure(0, weight=1)
        job_frame.columnconfigure(0, weight=1)

        self.job_list = tk.Listbox(
            job_frame,
            bg="#1e1e1e",
            fg="white",
            selectmode=tk.SINGLE,
            exportselection=False,
        )
        self.job_list.grid(row=0, column=0, sticky="nsew")

        job_scroll = tk.Scrollbar(job_frame, orient=tk.VERTICAL, command=self.job_list.yview)
        self.job_list.config(yscrollcommand=job_scroll.set)
        job_scroll.grid(row=0, column=1, sticky="ns")

        # manual selection tracking
        self.job_list.bind("<<ListboxSelect>>", self.on_job_select_manual)
        self.job_list.bind("<Key>", self.on_joblist_key)

        # Buttons: Add Invoice to Job + Add to Summary
        self.btn_add_invoice_to_job = tk.Button(
            frame,
            text="Add Invoice to Job",
            bg="#4caf50",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            command=self.add_invoice_to_job,
        )
        self.btn_add_invoice_to_job.grid(row=4, column=0, sticky="ew", padx=10, pady=(10, 5))

        self.btn_add_to_summary = tk.Button(
            frame,
            text="Add to Summary",
            bg="#0e639c",
            fg="white",
            font=("Segoe UI", 11, "bold"),
            command=self.add_to_summary,
        )
        self.btn_add_to_summary.grid(row=5, column=0, sticky="ew", padx=10, pady=(5, 15))

    # =========================================================
    # JOB LIST LOADING / BEHAVIOR
    # =========================================================
    def load_jobnames(self):
        self.jobnames = []
        self.job_list.delete(0, tk.END)
        if not os.path.exists(JOB_LIST_PATH):
            messagebox.showwarning(
                "Job List Missing",
                f"Job list file not found:\n{JOB_LIST_PATH}",
            )
            return
        try:
            with open(JOB_LIST_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    name = line.strip()
                    if name:
                        self.jobnames.append(name)
                        self.job_list.insert(tk.END, name)
        except Exception as e:
            messagebox.showerror("Error", f"Could not load job list:\n{e}")

    def on_job_select_manual(self, event):
        # User clicked or changed selection manually
        self.job_selection_auto = False

    def on_joblist_key(self, event):
        # Jump in job list by first letter (case-insensitive)
        ch = event.char.upper()
        if not ch.isalpha():
            return

        size = self.job_list.size()
        if size == 0:
            return

        # start search from next item after current selection
        cur = self.job_list.curselection()
        start = (cur[0] + 1) if cur else 0

        for i in range(start, size):
            text = self.job_list.get(i)
            if text.upper().startswith(ch):
                self.job_list.selection_clear(0, tk.END)
                self.job_list.selection_set(i)
                self.job_list.see(i)
                self.job_selection_auto = False  # user keyboard is "manual"
                return

        # wrap-around search
        for i in range(0, start):
            text = self.job_list.get(i)
            if text.upper().startswith(ch):
                self.job_list.selection_clear(0, tk.END)
                self.job_list.selection_set(i)
                self.job_list.see(i)
                self.job_selection_auto = False
                return

    # =========================================================
    # LOAD FOLDER
    # =========================================================
    def load_folder(self):
        folder = filedialog.askdirectory()
        if not folder:
            return

        self.current_folder = folder
        self.invoices = [
            os.path.join(folder, f)
            for f in os.listdir(folder)
            if f.lower().endswith(".pdf")
        ]
        self.invoice_info = {p: {"parsed": None, "jobname": None, "staged": False} for p in self.invoices}
        self.refresh_invoice_list()

        # set vendor module name based on folder name (and keep for processed)
        vendor_folder = os.path.basename(folder)
        self.vendor_module_name = self.derive_vendor_module(vendor_folder)
        self.vendor_label.config(text=f"Vendor: {vendor_folder}")

    # =========================== PATCH START ===========================
    def derive_vendor_module(self, vendor_folder: str) -> str:
        """
        Convert vendor folder names into Python-safe parser module names.
        Priority:
            1. Exact match in VENDOR_ALIASES
            2. Cleaned version: remove symbols → replace spaces with underscores
            3. Append _parser
        """
        raw = vendor_folder.lower().strip()

        # 1. Exact alias match
        if raw in VENDOR_ALIASES:
            return VENDOR_ALIASES[raw]

        # 2. Clean unexpected symbols (keep letters/numbers/underscores)
        cleaned = "".join(ch if ch.isalnum() else "_" for ch in raw)
        cleaned = "_".join(filter(None, cleaned.split("_")))  # collapse __

        return f"{cleaned}_parser"

    # =========================== PATCH END =============================

    def refresh_invoice_list(self):
        self.invoice_list.delete(0, tk.END)
        for p in self.invoices:
            idx = self.invoice_list.size()
            self.invoice_list.insert(tk.END, os.path.basename(p))
            info = self.invoice_info.get(p)
            if info and info.get("staged"):
                # light gray for staged invoices
                self.invoice_list.itemconfig(idx, fg="#808080")

    # =========================================================
    # WHEN A PDF IS SELECTED
    # =========================================================
    def on_invoice_select(self, event):
        if not self.invoice_list.curselection():
            return
        idx = self.invoice_list.curselection()[0]
        if idx < 0 or idx >= len(self.invoices):
            return

        pdf_path = self.invoices[idx]

        # preview PDF
        try:
            if self.pdf_document is not None:
                self.pdf_document.close()
        except Exception:
            pass

        try:
            self.pdf_document = fitz.open(pdf_path)
            self.displayed_page = 0
            self.render_pdf()
        except Exception as e:
            messagebox.showerror("Error", f"Cannot display PDF:\n{e}")

        # load parsed data (no raw text shown in UI)
        self.load_parsed_data(pdf_path)

    # =========================================================
    # RENDER PDF
    # =========================================================
    def render_pdf(self):
        if not self.pdf_document:
            return
        try:
            page = self.pdf_document.load_page(self.displayed_page)
        except Exception:
            return

        zoom = fitz.Matrix(self.zoom_level, self.zoom_level)
        pix = page.get_pixmap(matrix=zoom)

        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        self.pdf_image_ref = ImageTk.PhotoImage(img)  # keep reference

        self.pdf_canvas.delete("all")
        self.pdf_canvas.create_image(0, 0, anchor="nw", image=self.pdf_image_ref)
        self.pdf_canvas.config(scrollregion=self.pdf_canvas.bbox("all"))

    # =========================================================
    # ZOOM PDF
    # =========================================================
    def on_zoom(self, event):
        if event.delta > 0:
            self.zoom_level = min(self.zoom_level + 0.1, 6.0)
        else:
            self.zoom_level = max(self.zoom_level - 0.1, 0.4)
        self.render_pdf()

    # =========================================================
    # LOAD PARSED DATA (TOP ONLY)
    # =========================================================
    def load_parsed_data(self, pdf_path: str):
        self.parsed_output.delete("1.0", tk.END)
        self.last_parsed_invoice = None

        if not self.vendor_module_name:
            self.parsed_output.insert(tk.END, "No parser configured for this vendor.")
            return

        try:
            parser = __import__(self.vendor_module_name)
            result = parser.parse_invoice(pdf_path)
        except Exception as e:
            self.parsed_output.insert(tk.END, f"Parser error:\n{e}")
            return

        info = self.invoice_info.setdefault(pdf_path, {"parsed": None, "jobname": None, "staged": False})

        # result can be dict (single invoice) or list of dicts (multi-page)
        if isinstance(result, list):
            block_lines = []
            first = None
            for inv in result:
                if first is None:
                    first = inv
                block_lines.append(
                    "Vendor:         {vendor}\n"
                    "Invoice Number: {inv_no}\n"
                    "Jobname:        {job}\n"
                    "Date:           {date}\n"
                    "Total:          {total}\n".format(
                        vendor=inv.get("vendor", ""),
                        inv_no=inv.get("invoice_number", ""),
                        job=inv.get("jobname", ""),
                        date=inv.get("date", ""),
                        total=inv.get("total", ""),
                    )
                )
                block_lines.append("-" * 40)
            text = "\n".join(block_lines)
            self.parsed_output.insert(tk.END, text)
            if first:
                self.last_parsed_invoice = first
                info["parsed"] = first
                self.auto_highlight_job(first.get("jobname", ""))
        elif isinstance(result, dict):
            inv = result
            text = (
                "Vendor:         {vendor}\n"
                "Invoice Number: {inv_no}\n"
                "Jobname:        {job}\n"
                "Date:           {date}\n"
                "Total:          {total}\n".format(
                    vendor=inv.get("vendor", ""),
                    inv_no=inv.get("invoice_number", ""),
                    job=inv.get("jobname", ""),
                    date=inv.get("date", ""),
                    total=inv.get("total", ""),
                )
            )
            self.parsed_output.insert(tk.END, text)
            self.last_parsed_invoice = inv
            info["parsed"] = inv
            self.auto_highlight_job(inv.get("jobname", ""))
        else:
            self.parsed_output.insert(tk.END, str(result))

    def auto_highlight_job(self, jobname: str):
        """
        Highlight the most likely job in the job list based on parsed jobname.
        This is a HINT ONLY: the user must still manually select a job
        before Add Invoice to Job will run.
        """
        if not jobname:
            return
        target = jobname.lower().replace(" ", "")
        if not target:
            return

        best_index = None
        for idx in range(self.job_list.size()):
            text = self.job_list.get(idx)
            key = text.lower().replace(" ", "")
            if target in key or key in target:
                best_index = idx
                break

        if best_index is not None:
            self.job_selection_auto = True
            self.job_list.selection_clear(0, tk.END)
            self.job_list.selection_set(best_index)
            self.job_list.see(best_index)

    # =========================================================
    # PROCESS INVOICES (SPLIT + RENAME INTO processed/)
    # =========================================================
    def process_invoices(self):
        if not self.current_folder:
            messagebox.showerror("Error", "No folder selected.")
            return

        if not self.vendor_module_name:
            messagebox.showerror("Error", "No parser configured for this vendor.")
            return

        # Import vendor parser
        try:
            parser = __import__(self.vendor_module_name)
        except Exception as e:
            messagebox.showerror("Parser Error", f"Could not load parser:\n{e}")
            return

        processed_folder = os.path.join(self.current_folder, "processed")
        os.makedirs(processed_folder, exist_ok=True)

        total_created = 0

        for pdf_path in list(self.invoices):
            try:
                parsed_pages = parser.parse_invoice(pdf_path)
                if not isinstance(parsed_pages, list):
                    parsed_pages = [parsed_pages]

                doc = fitz.open(pdf_path)

                for inv in parsed_pages:
                    page_num = inv.get("page", 1) - 1
                    new_doc = fitz.open()
                    new_doc.insert_pdf(doc, from_page=page_num, to_page=page_num)

                    # ---------------------------
                    # Extract fields cleanly
                    # ---------------------------
                    invoice_num = inv.get("invoice_number", "") or "NOINV"
                    vendor = inv.get("vendor", "Vendor")
                    job = inv.get("jobname", "") or "Job"
                    date_str = inv.get("date", "") or ""
                    total_str = inv.get("total", "") or ""

                    # ---------------------------
                    # Clean for filename
                    # ---------------------------
                    vendor_clean = vendor.replace(" ", "")
                    job_clean = job.replace(" ", "").replace("/", "-")
                    date_clean = date_str.replace("/", "-")
                    invoice_clean = invoice_num.replace(" ", "")
                    total_clean = total_str.replace(",", "").replace(" ", "")

                    # Final filename ALWAYS includes invoice number
                    file_name = f"{vendor_clean}_{job_clean}_{date_clean}_{invoice_clean}_{total_clean}.pdf"

                    # Remove forbidden filename characters
                    bad_chars = ['"', "'", ":", "?", "*", "<", ">", "|"]
                    for ch in bad_chars:
                        file_name = file_name.replace(ch, "")

                    out_path = os.path.join(processed_folder, file_name)

                    new_doc.save(out_path)
                    new_doc.close()
                    total_created += 1

                doc.close()

            except Exception as e:
                messagebox.showwarning(
                    "Processing Error",
                    f"Error processing {os.path.basename(pdf_path)}:\n{e}",
                )

        # After a short delay, show processed folder contents
        self.after(200, lambda: self.show_processed_folder(processed_folder))

    def show_processed_folder(self, processed_folder: str):
        self.current_folder = processed_folder
        self.invoices = [
            os.path.join(processed_folder, f)
            for f in os.listdir(processed_folder)
            if f.lower().endswith(".pdf")
        ]
        # reset invoice info for processed invoices
        self.invoice_info = {
            p: {"parsed": None, "jobname": None, "staged": False}
            for p in self.invoices
        }
        self.refresh_invoice_list()
        # NOTE: Do NOT change vendor_module_name here; keep original vendor

    # =========================================================
    # ADD INVOICE TO JOB (STAGE + GRAY TEXT)
    # =========================================================
    def add_invoice_to_job(self):
        # Must have an invoice selected
        if not self.invoice_list.curselection():
            messagebox.showerror("Error", "Select an invoice first.")
            return
        inv_index = self.invoice_list.curselection()[0]
        if inv_index < 0 or inv_index >= len(self.invoices):
            return
        inv_path = self.invoices[inv_index]

        info = self.invoice_info.setdefault(inv_path, {"parsed": None, "jobname": None, "staged": False})
        if info.get("staged"):
            messagebox.showinfo("Already Added", "This invoice is already added to a job.")
            return

        # Job selection: user MUST manually choose (no auto-only)
        job_sel = self.job_list.curselection()
        if not job_sel or self.job_selection_auto:
            messagebox.showerror(
                "Job Required",
                "Please manually select a job from the list before adding the invoice to a job.",
            )
            return

        jobname = self.job_list.get(job_sel[0])

        # Ensure we have parsed data for this invoice
        if not info.get("parsed"):
            if not self.vendor_module_name:
                messagebox.showerror("Error", "No parser configured for this vendor.")
                return
            try:
                parser = __import__(self.vendor_module_name)
                result = parser.parse_invoice(inv_path)
                if isinstance(result, list):
                    inv = result[0] if result else {}
                elif isinstance(result, dict):
                    inv = result
                else:
                    inv = {}
                info["parsed"] = inv
                self.last_parsed_invoice = inv
            except Exception as e:
                messagebox.showerror("Parser Error", f"Could not parse invoice:\n{e}")
                return

        # Stage invoice to this job
        info["jobname"] = jobname
        info["staged"] = True

        # Turn invoice text in listbox light gray
        self.invoice_list.itemconfig(inv_index, fg="#808080")

        #messagebox.showinfo(
        #    "Staged",
        #    f"Invoice has been added to job '{jobname}'.\n"
        #    "You can now click 'Add to Summary' to write it to Excel and move the file."
        #)

    # =========================================================
    # SECRET CHEAT MODE — AUTO ADD MATCHING INVOICES (CTRL+R)
    # =========================================================
    def secret_auto_stage(self, event=None):
        """
        CTRL+R cheat code:
        Automatically stage all invoices where the parsed jobname
        exactly matches a job in the loaded job list.
        Case-insensitive, no fuzzy matching.
        """
        matches = 0

        # Build normalized lookup: "hamp" → "HAMP"
        job_lookup = {j.lower().strip(): j for j in self.jobnames}

        for idx, inv_path in enumerate(self.invoices):

            # Ensure invoice_info entry exists
            info = self.invoice_info.setdefault(
                inv_path,
                {"parsed": None, "jobname": None, "staged": False}
            )

            parsed = info.get("parsed")

            # Parse invoice now if needed
            if not parsed:
                try:
                    parser = __import__(self.vendor_module_name)
                    result = parser.parse_invoice(inv_path)
                    parsed = result[0] if isinstance(result, list) else result
                    info["parsed"] = parsed
                except Exception:
                    continue

            # Normalize jobname
            jn = (parsed.get("jobname") or "").strip().lower()

            # Exact match only (case ignored)
            if jn in job_lookup:
                jobname = job_lookup[jn]

                # Stage invoice
                info["jobname"] = jobname
                info["staged"] = True

                # Gray out in list
                self.invoice_list.itemconfig(idx, fg="#808080")

                matches += 1

        # Quiet status message
        self.parsed_output.insert(
            tk.END,
            f"\n\n[Secret Mode] Auto-staged {matches} invoice(s).\n"
        )

    # =========================================================
    # ADD TO SUMMARY (WRITE TO EXCEL + MOVE FILES)
    # =========================================================
    def add_to_summary(self):
        # --- FIX: Close any open PDF preview so Windows will let us move the file ---
        try:
            if self.pdf_document is not None:
                self.pdf_document.close()
                self.pdf_document = None
        except Exception:
            pass
        # ------------------------------------------------------------------------------

        # Collect all staged invoices
        staged_indices = []
        staged_paths = []
        for i, p in enumerate(self.invoices):
            info = self.invoice_info.get(p)
            if info and info.get("staged") and info.get("jobname"):
                staged_indices.append(i)
                staged_paths.append(p)

        if not staged_paths:
            messagebox.showerror(
                "Nothing Staged",
                "No invoices have been added to a job.\n"
                "Use 'Add Invoice to Job' first."
            )
            return

        # Ask user where to save the billing summary file
        excel_path = filedialog.asksaveasfilename(
            title="Save Billing Summary As",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile="Billing_summary.xlsx"
        )

        # If they cancel → abort
        if not excel_path:
            messagebox.showinfo("Cancelled", "Save operation cancelled. Nothing was added.")
            return

        # Ensure directory exists
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)

        # Load or create workbook
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()

        # ============================================
        # DUPLICATE CHECKING HELPER (NEW FUNCTION)
        # ============================================
        def invoice_exists_in_sheet(ws, date_str, vendor, invoice_number, total):
            """
            Returns True if this invoice already exists in the sheet
            (starting scan at row 18 downward).
            """
            row = 18
            while True:
                a = ws.cell(row=row, column=1).value  # date
                b = ws.cell(row=row, column=2).value  # vendor
                c = ws.cell(row=row, column=3).value  # invoice number
                d = ws.cell(row=row, column=4).value  # total

                # Hit empty row → not found
                if a is None and b is None and c is None and d is None:
                    return False

                if str(a) == str(date_str) and str(b) == str(vendor) \
                        and str(c) == str(invoice_number) and str(d) == str(total):
                    return True

                row += 1

        # Process each staged invoice
        processed_count = 0
        duplicate_count = 0

        for p in staged_paths:
            info = self.invoice_info.get(p, {})
            jobname = info.get("jobname")
            if not jobname:
                continue

            inv = info.get("parsed")
            if not inv:
                # Try to parse now if somehow missing
                if not self.vendor_module_name:
                    continue
                try:
                    parser = __import__(self.vendor_module_name)
                    result = parser.parse_invoice(p)
                    if isinstance(result, list):
                        inv = result[0] if result else {}
                    elif isinstance(result, dict):
                        inv = result
                    else:
                        inv = {}
                    info["parsed"] = inv
                except Exception:
                    continue

            date_str = inv.get("date", "")
            vendor = inv.get("vendor", "")
            invoice_number = inv.get("invoice_number", "")
            total = inv.get("total", "")

            # Get or create sheet for this jobname
            sheet_title = sanitize_sheet_title(jobname)
            if sheet_title in wb.sheetnames:
                ws = wb[sheet_title]
            else:
                ws = wb.create_sheet(title=sheet_title)

            # ============================================
            # VALIDATION — SKIP DUPLICATE ENTRIES
            # ============================================
            if invoice_exists_in_sheet(ws, date_str, vendor, invoice_number, total):
                duplicate_count += 1
                self.parsed_output.insert(
                    tk.END,
                    f"\nDuplicate skipped: Invoice {invoice_number} ({jobname})"
                )
                continue

            # Find first empty row at or below row 18
            row = 18
            while ws.cell(row=row, column=1).value not in (None, ""):
                row += 1

            ws.cell(row=row, column=1, value=date_str)
            ws.cell(row=row, column=2, value=vendor)
            ws.cell(row=row, column=3, value=invoice_number)
            ws.cell(row=row, column=4, value=total)

            # Move processed file into ROOT/Billing_Summary/<jobname>/invoices
            job_folder_safe = sanitize_sheet_title(jobname)
            invoices_folder = os.path.join(BILLING_SUMMARY_ROOT, job_folder_safe, "invoices")
            safe_move(p, invoices_folder)

            processed_count += 1

        # Remove default sheet if it's unused
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            std = wb["Sheet"]
            if all(std.cell(row=r, column=c).value is None for r in range(1, 10) for c in range(1, 5)):
                wb.remove(std)

        wb.save(excel_path)

        # Remove staged invoices from list (ONLY those written)
        for i in sorted(staged_indices, reverse=True):
            path = self.invoices[i]
            info = self.invoice_info.get(path)
            if info and info.get("staged") and info.get("jobname"):
                if path in staged_paths:  # already validated
                    self.invoice_list.delete(i)
                    del self.invoices[i]
                    self.invoice_info.pop(path, None)

        # Report results
        self.parsed_output.insert(
            tk.END,
            f"\n\n{processed_count} invoice(s) added to summary."
        )
        if duplicate_count > 0:
            self.parsed_output.insert(
                tk.END,
                f"\n{duplicate_count} duplicate(s) skipped."
            )

        messagebox.showinfo(
            "Success",
            f"{processed_count} invoice(s) added.\n{duplicate_count} duplicate(s) skipped."
        )

        # Reset flags for next round
        self.last_parsed_invoice = None
        self.job_selection_auto = False


# =========================================================
# PROGRAM ENTRY POINT
# =========================================================
if __name__ == "__main__":
    app = InvoiceSorterUI()
    app.mainloop()
